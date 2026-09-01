#!/usr/bin/env python3
"""Build one downloadable workbook per grade from the generated markdown plans.

Each file carries every plan for that grade plus the shared red-day sheet, and
is written both here and into site/downloads/ so the website can serve it.

Run after the two plan generators:  python3 build-xlsx-per-grade.py
"""
import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
SITE = os.path.join(HERE, os.pardir, 'site', 'downloads')

GRADES = {
    6: ('Grade06_Mathematics_Annual_Plan_2026-2027.xlsx', [
        ('08-cambridge-class-grade6-mathematics.md', 'G6 Mathematics 6h',
         'ANNUAL PLAN — MATHEMATICS, GRADE 6 (CAMBRIDGE CLASS)',
         '6 hours per week · 204 hours · national KTP (170 h) with a 34-hour Cambridge Lower Secondary Stage 7 layer'),
    ]),
    7: ('Grade07_Mathematics_Annual_Plan_2026-2027.xlsx', [
        ('09-special-class-mathematics-grade7.md', 'G7 Mathematics 5h',
         'ANNUAL PLAN — MATHEMATICS, GRADE 7 (SPECIAL CLASS)',
         '5 hours per week · 170 hours · national KTP enriched with Cambridge Lower Secondary Stage 8'),
    ]),
    8: ('Grade08_Annual_Plans_2026-2027.xlsx', [
        ('01-special-class-algebra-grade8.md', 'G8 Algebra 3h',
         'ANNUAL PLAN — ALGEBRA, GRADE 8 (SPECIAL CLASS)',
         '3 hours per week · 102 hours · national KTP enriched with Cambridge Lower Secondary Stage 9'),
        ('02-special-class-geometry-grade8.md', 'G8 Geometry 2h',
         'ANNUAL PLAN — GEOMETRY, GRADE 8 (SPECIAL CLASS)',
         '2 hours per week · 68 hours · national KTP enriched with Cambridge Lower Secondary Stage 9'),
        ('03-cambridge-class-grade8-mathematics.md', 'G8 Cambridge 6h',
         'ANNUAL PLAN — MATHEMATICS, GRADE 8 (CAMBRIDGE CLASS)',
         '6 hours per week · 204 hours · Cambridge Stage 9 backbone with the Uzbekistan Grade 8 programme as depth'),
    ]),
    9: ('Grade09_Annual_Plans_2026-2027.xlsx', [
        ('10-special-class-algebra-grade9.md', 'G9 Algebra 3h',
         'ANNUAL PLAN — ALGEBRA, GRADE 9 (SPECIAL CLASS)',
         '3 hours per week · 102 hours · national KTP enriched with Cambridge IGCSE Extended, chapters 1-12'),
        ('11-special-class-geometry-grade9.md', 'G9 Geometry 2h',
         'ANNUAL PLAN — GEOMETRY, GRADE 9 (SPECIAL CLASS)',
         '2 hours per week · 68 hours · national KTP enriched with Cambridge IGCSE Extended, chapters 1-12'),
    ]),
    10: ('Grade10_Annual_Plans_2026-2027.xlsx', [
        ('04-special-class-algebra-grade10.md', 'G10 Algebra 3h',
         'ANNUAL PLAN — ALGEBRA, GRADE 10 (SPECIAL CLASS)',
         '3 hours per week · 102 hours · national KTP enriched with Cambridge IGCSE Extended and AS Pure Mathematics 1'),
        ('05-special-class-geometry-grade10.md', 'G10 Geometry 2h',
         'ANNUAL PLAN — GEOMETRY, GRADE 10 (SPECIAL CLASS)',
         '2 hours per week · 68 hours · national KTP enriched with Cambridge IGCSE Extended and AS Pure Mathematics 1'),
    ]),
    11: ('Grade11_Annual_Plans_2026-2027.xlsx', [
        ('06-special-class-algebra-grade11.md', 'G11 Algebra 3h',
         'ANNUAL PLAN — ALGEBRA AND CALCULUS, GRADE 11 (SPECIAL CLASS)',
         '3 hours per week · 102 hours · national KTP enriched with Cambridge AS & A Level Pure Mathematics 1 and 2'),
        ('07-special-class-geometry-grade11.md', 'G11 Geometry 2h',
         'ANNUAL PLAN — GEOMETRY, GRADE 11 (SPECIAL CLASS)',
         '2 hours per week · 68 hours · national KTP enriched with Cambridge IGCSE Extended and A Level Pure Mathematics 2'),
    ]),
}

BRAND   = '0E5C63'
TINT    = 'E4F0F0'
QUARTER = '18272B'
FILLS = {'Test': 'F8E9E4', 'CAM': 'E6F1E8', 'R': 'F7EEDA'}
FONTS = {'Test': 'A34430', 'CAM': '3C7A50', 'R': 'B0801F'}

thin = Side(style='thin', color='D8D5CC')
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def read(md):
    """Pull (quarter heading | row) records out of a generated plan."""
    recs = []
    for line in open(os.path.join(HERE, md), encoding='utf-8'):
        q = re.match(r'##\s+(I{1,3}V?)\s+QUARTER\s+\((\d+)\s+hours\)', line)
        if q:
            recs.append(('Q', '%s QUARTER (%s hours)' % (q.group(1), q.group(2))))
            continue
        m = re.match(r'\|\s*(\d+)\s*\|\s*(.+?)\s*\|\s*(\d+)\s*\|\s*(.*?)\s*\|\s*(\S*)\s*\|', line)
        if m:
            recs.append(('L', (int(m.group(1)), m.group(2), int(m.group(3)),
                               m.group(4), m.group(5))))
    return recs


def build(sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for md, tab, title, sub in sheets:
        ws = wb.create_sheet(tab)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=13, color=BRAND)
        ws['A2'] = sub
        ws['A2'].font = Font(size=9, italic=True, color='5F7076')
        ws.append([])
        ws.append(['No.', 'Lesson topic', 'Hours', 'Cambridge link / note', 'Flag'])
        for c in ws[4]:
            c.font = Font(bold=True, size=10, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=BRAND)
            c.alignment = Alignment(vertical='center')
            c.border = box

        for kind, payload in read(md):
            if kind == 'Q':
                ws.append([payload])
                r = ws.max_row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
                cell = ws.cell(row=r, column=1)
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor=QUARTER)
                cell.alignment = Alignment(vertical='center')
            else:
                ws.append(list(payload))
                r = ws.max_row
                flag = payload[4]
                for col in range(1, 6):
                    cell = ws.cell(row=r, column=col)
                    cell.border = box
                    cell.alignment = Alignment(vertical='top', wrap_text=(col in (2, 4)))
                    cell.font = Font(size=10)
                if flag in FILLS:
                    for col in range(1, 6):
                        ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=FILLS[flag])
                    ws.cell(row=r, column=5).font = Font(size=10, bold=True, color=FONTS[flag])

        for col, w in zip('ABCDE', (6, 62, 7, 52, 10)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A5'

    # ---- red days sheet ----
    ws = wb.create_sheet('Red days & reserve')
    ws['A1'] = 'RED DAYS — UZBEKISTAN PUBLIC HOLIDAYS, 2026–2027'
    ws['A1'].font = Font(bold=True, size=13, color=BRAND)
    ws.append([])
    ws.append(['Date', 'Holiday', 'Weekday', 'Effect on the plan'])
    for c in ws[3]:
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=BRAND)
        c.border = box
    DAYS = [
        ('1 September', 'Independence Day', 'Tue', 'Falls before the first school day (2 September) — no lesson lost.'),
        ('1 October', 'Day of Teachers and Mentors', 'Thu', '1 school day lost — Quarter I.'),
        ('8 December', 'Constitution Day', 'Tue', '1 school day lost — Quarter II.'),
        ('1 January', 'New Year', 'Fri', 'Falls inside the winter holiday — no lesson lost.'),
        ('14 January', 'Day of Defenders of the Motherland', 'Thu', '1 school day lost — Quarter III.'),
        ('8 March', "International Women's Day", 'Mon', '1 school day lost — Quarter III.'),
        ('21 March', 'Navruz', 'Sun', 'Normally inside the spring holiday; the day off in lieu may cost 1 school day.'),
        ('Ramazon Hayit (movable)', 'Eid al-Fitr', '—', 'Approx. late March 2027; usually adjoins the spring holiday. Confirm annually.'),
        ('9 May', 'Day of Remembrance and Honour', 'Sun', 'Day off in lieu on Monday 10 May — 1 school day lost, Quarter IV.'),
        ('Qurbon Hayit (movable)', 'Eid al-Adha', '—', 'Approx. late May 2027; may fall after the last lesson. Confirm annually.'),
    ]
    for d in DAYS:
        ws.append(list(d))
        for col in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = box
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 4))
    ws.append([])
    ws.append(['ABSORPTION PROTOCOL'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, color=BRAND)
    for step in [
        '1. At the start of each quarter, count the red days falling on this class’s mathematics days.',
        '2. Remove exactly that many R lessons from the same quarter, working from the end backwards.',
        '3. Never remove a Test lesson and never remove a lesson that introduces new content.',
        '4. If a quarter has too few R lessons, merge two consecutive lessons with the same title.',
        '5. Record every adjustment so the quarter still shows the official hour total.',
    ]:
        ws.append([step])
        ws.cell(row=ws.max_row, column=1).font = Font(size=10)
    for col, w in zip('ABCD', (24, 34, 10, 66)):
        ws.column_dimensions[col].width = w

    #print('wrote', os.path.basename(OUT), '|', len(wb.sheetnames), 'sheets:', ', '.join(wb.sheetnames))
    return wb


if __name__ == '__main__':
    if not os.path.isdir(SITE):
        os.makedirs(SITE)
    for g in sorted(GRADES):
        fname, sheets = GRADES[g]
        wb = build(sheets)
        for path in (os.path.join(HERE, fname), os.path.join(SITE, fname)):
            wb.save(path)
        print('%-52s %d plan sheet(s) + red days' % (fname, len(sheets)))
